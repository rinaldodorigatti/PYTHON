import openpyxl

wb = openpyxl.load_workbook('FILES/quatre.xlsx')
print(wb.sheetnames)

sheet = wb['Les ouates']

PRICE_UPDATES = {'NAME': 'Sara', 'PHONE': '0216359251', 'ADDRESS': 'Chemin de la Prairie 61', 'CITY': 'Renens'}
PRICE_UPDATES_02 = {'NAME': 'Rinaldo', 'PHONE': '0216359255', 'ADDRESS': 'Chemin de la Prairie 61', 'CITY': 'Lausanne'}
PRICE_UPDATES_03 = {'NAME': 'Lucie', 'PHONE': '0216359258', 'ADDRESS': 'Chemin de la Prairie 61', 'CITY': 'Zurich'}
PRICE_UPDATES_04 = {'NAME': 'Diego', 'PHONE': '0216359260', 'ADDRESS': 'Chemin de la Prairie 61', 'CITY': 'Sion'}


for k in range(1, sheet.max_column + 1):
    produceName = sheet.cell(row=1, column=k).value
    if produceName in PRICE_UPDATES:
        sheet.cell(row=2, column=k).value = PRICE_UPDATES[produceName]
        sheet.cell(row=3, column=k).value = PRICE_UPDATES_02[produceName]
        sheet.cell(row=4, column=k).value = PRICE_UPDATES_03[produceName]
        sheet.cell(row=5, column=k).value = PRICE_UPDATES_04[produceName]
wb.save('FILES/quatre.xlsx')
