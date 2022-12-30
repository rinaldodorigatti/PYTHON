import json
from openpyxl import load_workbook

workbook = load_workbook(filename='FILES/copy_deux.xlsx')
sheet = workbook.active

products = {}

for row in sheet.iter_rows(min_row=2, min_col=1, max_col=4, values_only=True):
    product_id = row[0]
    product = {
        "date": row[1],
        "what": row[2],
        "quantity": row[3]
    }
    products[product_id] = product

print(json.dumps(products))
