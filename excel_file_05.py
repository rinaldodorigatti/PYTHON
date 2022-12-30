import openpyxl
from openpyxl.styles import Font
from openpyxl.styles.borders import Border, Side


wb = openpyxl.load_workbook('FILES/quatre.xlsx')
print(wb.sheetnames)

sheet = wb['Les ouates']

PRICE_UPDATES_ALL = (
    {'NAME': 'Sara', 'PHONE': '0216359251', 'ADDRESS': 'Chemin de la Prairie 61', 'CITY': 'Renens'},
    {'NAME': 'Rinaldo', 'PHONE': '0216359255', 'ADDRESS': 'Chemin de la Prairie 61', 'CITY': 'Lausanne'},
    {'NAME': 'Lucie', 'PHONE': '0216359258', 'ADDRESS': 'Chemin de la Prairie 61', 'CITY': 'Zurich'},
    {'NAME': 'Diego', 'PHONE': '0216359260', 'ADDRESS': 'Chemin de la Prairie 61', 'CITY': 'Sion'}
)

t = ['NAME', 'PHONE', 'ADDRESS', 'CITY']

for r in PRICE_UPDATES_ALL:
    for g in r.keys():
        print("%10s" % g, end=' ')
        if g in t:
            print("%30s" % r[g])
    print()

columns = []
for k in range(1, sheet.max_column + 1):
    columns.append(sheet.cell(row=1, column=k).value)

# print(columns)

for k in range(1, sheet.max_column + 1):
    produceName = sheet.cell(row=1, column=k).value

price_updates_keys = []
price_updates_values = []
for lll in range(0, len(PRICE_UPDATES_ALL)):
    test = PRICE_UPDATES_ALL[lll].keys()
    test2 = PRICE_UPDATES_ALL[lll].values()
    price_updates_keys.append(test)
    price_updates_values.append(test2)

PRICE_UPDATES_ONE = (
    {'NAME': 'Sara', 'PHONE': '0216359251', 'ADDRESS': 'Chemin de la Prairie 61', 'CITY': 'Renens', 'PAYE': 2000},
    {'NAME': 'Rinaldo', 'PHONE': '0216359255', 'ADDRESS': 'Chemin de la Prairie 61', 'CITY': 'Lausanne', 'PAYE': 3500},
    {'NAME': 'Lucie', 'PHONE': '0216359258', 'ADDRESS': 'Chemin de la Prairie 61', 'CITY': 'Zurich', 'PAYE': 2850},
    {'NAME': 'Diego', 'PHONE': '0216359260', 'ADDRESS': 'Chemin de la Prairie 61', 'CITY': 'Sion', 'PAYE': 1300}
)

# print(PRICE_UPDATES_ONE[0].get('NAME'))
# print(PRICE_UPDATES_ONE[0].keys())

boldFont = Font(bold=True)
sheet['A1'].font = boldFont
sheet['B1'].font = boldFont
sheet['C1'].font = boldFont
sheet['D1'].font = boldFont
sheet['E1'].font = boldFont
sheet['D7'] = 'Total'
sheet['D7'].font = boldFont
sheet['E7'] = '=SUM(E2:E6)'

sheet.row_dimensions[1].height = 70
sheet.column_dimensions['C'].width = 50

sheet.merge_cells('A9:B9')
# sheet.unmerge_cells('A9:B9')
sheet['A9'] = 'Test de fichier'

borderStyle = openpyxl.styles.Side(style='medium', color='000000')
sheet['A1'].border = openpyxl.styles.Border(left=borderStyle, right=borderStyle, top=borderStyle, bottom=borderStyle)
sheet['B1'].border = openpyxl.styles.Border(left=borderStyle, right=borderStyle, top=borderStyle, bottom=borderStyle)
sheet['C1'].border = openpyxl.styles.Border(left=borderStyle, right=borderStyle, top=borderStyle, bottom=borderStyle)
sheet['D1'].border = openpyxl.styles.Border(left=borderStyle, right=borderStyle, top=borderStyle, bottom=borderStyle)
sheet['E1'].border = openpyxl.styles.Border(left=borderStyle, right=borderStyle, top=borderStyle, bottom=borderStyle)

# borderStyle11 = openpyxl.styles.Side(style='thin', color='000000')
borderStyle11 = Border(left=Side(style='thin'), bottom=Side(style='thin'),
                       top=Side(style='thin'), right=Side(style='thin'))

for col in range(1, sheet.max_column + 1):
    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row=row, column=col)
        cell.border = borderStyle11


"""
myStyle = openpyxl.styles.NamedStyle(name = 'my_style')
myStyle.font = openpyxl.styles.Font(name = 'Courier', size = 12, color = 'FF0000')
myStyle.fill = openpyxl.styles.PatternFill(patternType = 'solid', fgColor = 'FFFF55')
borderStyle = openpyxl.styles.Side(style = 'dashDot', color = 'FF00FF')
myStyle.border = openpyxl.styles.Border(left = borderStyle, right = borderStyle, top = borderStyle, bottom = borderStyle)

Les styles des bordures peuvent être :
'dashDotDot', 'mediumDashDotDot', 'medium'
'double', 'dotted', 'hair', 'mediumDashDot'
'mediumDashed', 'thick', 'dashed', 'dashDot'
'thin', 'slantDashDot'
"""

for k in range(1, sheet.max_column + 1):
    produceName = sheet.cell(row=1, column=k).value
    if produceName in PRICE_UPDATES_ONE[0].keys():
        sheet.cell(row=2, column=k).value = PRICE_UPDATES_ONE[0].get(produceName)
        sheet.cell(row=3, column=k).value = PRICE_UPDATES_ONE[1].get(produceName)
        sheet.cell(row=4, column=k).value = PRICE_UPDATES_ONE[2].get(produceName)
        sheet.cell(row=5, column=k).value = PRICE_UPDATES_ONE[3].get(produceName)
        sheet.cell(row=6, column=k).value = PRICE_UPDATES_ONE[3].get(produceName)

wb.save('FILES/quatre.xlsx')
