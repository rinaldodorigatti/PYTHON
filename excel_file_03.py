import openpyxl

wb = openpyxl.load_workbook('FILES/trois.xlsx')
print(wb.sheetnames)
"""
wb.create_sheet(index=1, title='Les plans')
print(wb.sheetnames)
wb.create_sheet(index=2, title='Les ouates')
print(wb.sheetnames)
"""
sheet = wb['Les ouates']
"""
sheet['A1'] = 'NAME'
sheet['B1'] = 'PHONE'
sheet['C1'] = 'ADDRESS'
sheet['D1'] = 'CITY'
"""

PRICE_UPDATES = {'NAME': 'Sara', 'PHONE': '0216359251', 'ADDRESS': 'Chemin de la Prairie 61', 'CITY': 'Renens'}

for k in range(1, sheet.max_row + 1):
    produceName = sheet.cell(row=k, column=1).value
    if produceName in PRICE_UPDATES:
        sheet.cell(row=k, column=2).value = PRICE_UPDATES[produceName]
wb.save('FILES/trois.xlsx')
