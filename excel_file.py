import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string


wb = openpyxl.load_workbook('FILES/un.xlsx')
print(type(wb))
print(wb.sheetnames)
sheet = wb['Sheet1']  # get sheet from workbook
print(sheet.title)
anotherSheet = wb.active
print(anotherSheet)
print(sheet['A1'].value)
cellB = sheet['B1']
print(cellB.value)
print('Row %s column %s is : %s' % (cellB.row, cellB.column, cellB.value))
print("row1 col2 :", sheet.cell(row=1, column=2).value)

print("-" * 35)
print('%-5s %-20s %-10s' % ('NUM', 'DESC', 'BUY'))
print("-" * 35)
for i in range(1, 8, 1):
    print('%-5s %-20s %-10s' % (i, sheet.cell(row=i, column=2).value, sheet.cell(row=i, column=3).value))
print("-" * 35)

print("max_row :", sheet.max_row, "\tmax_column :", sheet.max_column)

print("get_column_letter 1      :", get_column_letter(1))
print("get_column_letter 2      :", get_column_letter(2))
print("get_column_letter max    :", get_column_letter(sheet.max_column))
print("column_index_from_string :", column_index_from_string('A'))
print("column_index_from_string :", column_index_from_string('AA'))
print("column_index_from_string :", column_index_from_string(get_column_letter(sheet.max_column)))

wb1 = openpyxl.load_workbook('FILES/un.xlsx')
sheet1 = wb1['Sheet1']

print("-" * 35)
for rowOfCellObjects in sheet1['A1':'C3']:
    for cellObj in rowOfCellObjects:
        print(cellObj.coordinate, cellObj.value)
    print('----- END OF ROW -----')


wb2 = openpyxl.load_workbook('FILES/un.xlsx')
sheet2 = wb2.active
# laliste = list(sheet2.columns)[1]
# print(laliste)

print("-" * 35)
for cells in list(sheet2.columns)[1]:
    print(cells.value)
print("-" * 35)

print(wb2.sheetnames)
sheet2.title = 'Spam Bacon Eggs Sheet'
print(wb2.sheetnames)
