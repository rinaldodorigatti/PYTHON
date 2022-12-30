import openpyxl

print("Opening Workbook...")

wb = openpyxl.load_workbook('FILES/deux.xlsx')
sheet = wb['Les Fruits']
countyData = {}

print("Reading row...")

for row in range(2, sheet.max_row + 1):
    state = sheet['A' + str(row)].value
    county = sheet['B' + str(row)].value
    pop = sheet['C' + str(row)].value
    print(state, county, pop)

print(wb.sheetnames)
sheet = wb.active
print(sheet.title)
sheet.title = 'Les pommes'
print(sheet.title)
wb.save("copy_deux.xlsx")
