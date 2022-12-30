import openpyxl
from datetime import datetime

wb = openpyxl.Workbook()
sheet = wb.active
sheet['A1'] = 'Software IBM testing'
sheet.cell(row=2, column=1).value = 'Oracle DB'
now = datetime.now()
date_time = now.strftime('%d.%m.%Y')
sheet.title = date_time
sheetname = "Result " + date_time
wb.create_sheet(index=1, title=sheetname)

sheet1 = wb['Result 27.08.2022']
data = [('Emp Id', 'Emp Name', 'Designation'),
        (1, 'XYZ', 'Manager'),
        (2, 'ABC', 'Consultant')]

for i in data:
    sheet1.append(i)


def print_up_list(up_list, up_meta_info, base):
    from openpyxl.workbook import Workbook
    from openpyxl.worksheet import worksheet

    filename = 'FILES/UP_print.xlsx'
    wb1 = Workbook(encoding='mac_roman')
    ws = sheet.title(wb1, title='UP_list')
    ws.freeze_panes = 'A2'
    header = ['#',
              'UP name',
              'unit',
              'country',
              'infrastructure']
    for ie in range(6):
        header.append('Category ' + str(ie))
    ws.append(header)

    for ir in range(len(up_list)):
        up = up_list[ir]
        line = [ir + base,
                up_list[ir],
                up_meta_info[up]['unit'],
                up_meta_info[up]['Country'],
                up_meta_info[up]['Infrastructure']]
        for j in range(6):
            try:
                line.append(up_meta_info[up]['Category type'][j])
            except IndexError:
                break
        ws.append(line)
    print ('saving in excel sheet named: ' + filename)
    wb.add_sheet(ws)
    wb.save(filename)


data1 = [(1, 'Emp Name', 'Designation'), (2, 'XYZ', 'Manager'), (3, 'ABC', 'Consultant')]

for col in range(1, sheet1.max_column + 1):
    for row in range(2, sheet1.max_row + 1):
        sheet1.cell(row=, column=)

wb.save('FILES/new_file.xlsx')
